import csv
import random
from openpyxl.workbook import Workbook

# 定义可能的班级、课程、教师、课时、课程类型、教室类型和开始周
grades = [21, 22, 23, 24]
majors = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H','I']
class_numbers = [1, 2, 3,4,5]
public_courses = [f'P_{i:02d}' for i in range(1, 10)]
professional_courses = [f'M_{i:02d}' for i in range(1, 70)]
public_teachers = [f'PT_{i:02d}' for i in range(1, 30)]
professional_teachers = [f'MT_{i:02d}' for i in range(1, 70)]
class_hours = [8, 12, 16, 20, 24, 32, 36]
course_types = ['m', 'p']
start_week = 1

classroom_types = []
classroom_types.append('lc,mc')
classroom_types.append('mc')
classroom_types.append('bc')

classroom_types1 = []
classroom_types1.append('ll,ml')
classroom_types1.append('ml')

def gen():
    for i in range(9):

        min_num = 3
        max_num = 8

        # 随机确定要选择的元素数量
        numbers = [3, 4, 5, 6, 7, 8,9]
        weights = [0.15, 0.2, 0.2,0.2,0.15,0.05,0.05]
        num_to_choose = random.choices(numbers, weights=weights)[0]

        # 随机选择不重复的元素
        selected_majors = random.sample(majors, num_to_choose)
        print(selected_majors)

        numbers = [1, 2, 3, 4, 5]
        weights = [0.2,0.4,0.2,0.1,0.1]
        class_numbers = random.choices(numbers, weights=weights)[0]

        # 生成班级列表
        classes = []
        for major in selected_majors:
            for grade in grades:
                numbers = [1, 2, 3, 4, 5]
                weights = [0.2, 0.4, 0.2, 0.1, 0.1]
                class_numbers = random.choices(numbers, weights=weights)[0]
                for class_number in range(class_numbers):
                    classes.append(f'{grade}_{major}_{class_number:02d}')

        print(len(classes))

        # 初始化班级选择次数记录字典
        class_selection_count = {cls: 0 for cls in classes}
        # 初始化教师选择次数记录字典
        teacher_selection_count = {teacher: 0 for teacher in public_teachers + professional_teachers}

        # 生成课程计划数据
        data = []
        id_counter = 1
        for _ in range(len(classes)*4):
            # 随机选择班级数量
            numbers = [1, 2, 3]
            weights = [0.5, 0.4, 0.1]
            num_classes = random.choices(numbers, weights=weights)[0]

            selected_classes = []
            available_classes = classes.copy()
            while len(selected_classes) < num_classes and available_classes:
                cls = random.choice(available_classes)
                if class_selection_count[cls] < 10:
                    selected_classes.append(cls)
                    class_selection_count[cls] += 1
                available_classes.remove(cls)

            class_str = ','.join(selected_classes)

            # 随机选择课程类型
            numbers = [0, 1]
            weights = [0.7, 0.3]
            course_type = course_types[random.choices(numbers, weights=weights)[0]]

            if course_type == 'p':
                available_teachers = [t for t in public_teachers if teacher_selection_count[t] < 5]
                if not available_teachers:
                    continue
                teacher = random.choice(available_teachers)
                course = random.choice(public_courses)
            else:
                available_teachers = [t for t in professional_teachers if teacher_selection_count[t] < 5]
                if not available_teachers:
                    continue
                teacher = random.choice(available_teachers)
                course = random.choice(professional_courses)

            teacher_selection_count[teacher] += 1

            # 随机选择课时
            numbers = [8, 12, 16, 20, 24, 32, 36]
            weights = [1.5, 1.5, 1.5, 1, 1, 1, 0.05]
            hours = random.choices(numbers, weights=weights)[0]

            # 根据班级数量选择教室类型
            classroom = classroom_types[num_classes - 1]

            # 如果是专业课且有实验课
            if course_type == 'm' and random.random() < 0.5 and num_classes < 3:
                theory_course = course
                lab_course = theory_course.replace('M', 'L')
                # 实验课课时大概是理论课的 1/3 - 1/2
                lab_hours = random.randint(hours // 3, hours // 2)
                # 理论课数据
                data.append([id_counter, class_str, theory_course, teacher, hours, course_type, classroom, start_week])
                id_counter += 1
                classroom = classroom_types1[num_classes - 1]
                # 实验课数据
                data.append([id_counter, class_str, lab_course, teacher, lab_hours, course_type, classroom,
                             start_week])
                id_counter += 1
            else:
                data.append([id_counter, class_str, course, teacher, hours, course_type, classroom, start_week])
                id_counter += 1

        # 创建一个新的工作簿和工作表
        wb = Workbook()
        ws = wb.active
        ws.title = 'CT_test'

        # 写入表头
        ws.append(['id', 'class', 'course', 'teacher', 'time', 'type', 'cr_type', 'start_week'])

        # 写入数据
        for row in data:
            ws.append(row)

        # 提取所有班级
        all_classes = set()
        for row in data:
            class_str = row[1]
            classes_in_row = class_str.split(',')
            for c in classes_in_row:
                all_classes.add(c)

        # 新建一个工作表来存储班级信息
        class_sheet = wb.create_sheet('CL')
        class_sheet.append(['id', 'name'])
        class_id = 1
        for class_name in sorted(all_classes):
            class_sheet.append([class_id, class_name])
            class_id += 1

        # 提取所有老师
        all_teachers = set()
        for row in data:
            teacher = row[3]
            all_teachers.add(teacher)

        # 新建一个工作表来存储老师信息
        teacher_sheet = wb.create_sheet('M_T')
        teacher_sheet.append(['id', 'name', 'like1', 'like3'])
        teacher_id = 1
        for teacher_name in sorted(all_teachers):
            numbers = [-1, 0, 1]
            weights = [0.3, 0.4, 0.3]
            like1 = ','.join(str(random.choices(numbers, weights=weights)[0]) for _ in range(5))
            weights = [0.3, 0.4, 0.3]
            like3 = ','.join(str(random.choices(numbers, weights=weights)[0]) for _ in range(1))
            teacher_sheet.append([teacher_id, teacher_name, like1, like3])
            teacher_id += 1

        # 提取所有课程
        all_courses = set()
        for row in data:
            course = row[2]
            all_courses.add(course)

        # 新建一个工作表来存储课程信息
        course_sheet = wb.create_sheet('M_C')
        course_sheet.append(['id', 'name', 'like'])
        course_id = 1
        for course_name in sorted(all_courses):
            numbers = [-1, 0, 1]
            weights = [0.2, 0.6, 0.2]
            like = ','.join(str(random.choices(numbers, weights=weights)[0]) for _ in range(5))
            course_sheet.append([course_id, course_name, like])
            course_id += 1

        # 保存工作簿
        name = 'data' + str(i + 2) + '.xlsx'
        wb.save(name)

        print("课程计划数据已生成并保存到 " + name + " 文件中。")
